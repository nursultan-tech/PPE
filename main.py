import os
import numpy as np
import face_recognition
import cv2
from datetime import datetime, timedelta
from ultralytics import YOLO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk

# Конфигурация для загрузки модели YOLO
class CFG:
    WEIGHTS = 'C:\\Users\\Acer\\OneDrive\\Рабочий стол\\PPE\\model\\yolo-ppe-100-epochs\\runs\\detect\\yolov9e_ppe_css_50_epochs\\weights\\best.pt'
    CONFIDENCE = 0.35
    MIN_BOX_SIZE = 20
    CLASSES_TO_DETECT = [0, 2, 4, 7]

# Загрузка модели
model = YOLO(CFG.WEIGHTS)
model.conf = CFG.CONFIDENCE

# Функция для загрузки лиц из папки dataset
def load_known_faces(dataset_path='dataset'):
    known_face_encodings = []
    known_face_names = []

    for filename in os.listdir(dataset_path):
        if filename.endswith(".jpg"):
            image_path = os.path.join(dataset_path, filename)
            image = face_recognition.load_image_file(image_path)
            encoding = face_recognition.face_encodings(image)[0]
            known_face_encodings.append(encoding)
            name = os.path.splitext(filename)[0].replace('_', ' ')
            known_face_names.append(name)

    return known_face_encodings, known_face_names

# Загрузка и кодирование лиц
known_face_encodings, known_faces_names = load_known_faces()

# Последние метки времени для каждого студента
last_logged_time = {name: None for name in known_faces_names}
last_full_record_time = datetime.now()

# Интервал для полного логирования
full_log_interval = timedelta(seconds=5)

# Создание файла для записи посещаемости
excel_file = 'output.xlsx'
wb = Workbook()
ws = wb.active
ws.title = "Attendance Log"

# Установка заголовков
headers = ["Time", "Worker", "Hardhat", "Safety Vest"]
ws.append(headers)

for cell in ws["1:1"]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

# Функция для открытия Excel файла
def open_excel_file():
    if os.path.exists(excel_file):
        os.system(f'start excel "{excel_file}"')
    else:
        messagebox.showerror("Error", f"File {excel_file} not found!")

# Функция для обновления видеопотока
def update_frame():
    global last_full_record_time, full_log_interval, known_faces_names
    
    ret, frame = video_capture.read()
    if not ret:
        print("Failed to grab frame")
        return

    # Поиск лиц в текущем кадре
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
    face_locations = face_recognition.face_locations(rgb_small_frame)
    face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
    face_names = []

    if len(face_locations) > 0:
        # Найти наиболее явно видимое лицо (наибольший размер)
        largest_face_index = np.argmax([bottom - top for top, right, bottom, left in face_locations])
        largest_face_location = face_locations[largest_face_index]
        largest_face_encoding = face_encodings[largest_face_index]

        matches = face_recognition.compare_faces(known_face_encodings, largest_face_encoding, tolerance=0.4)
        face_distances = face_recognition.face_distance(known_face_encodings, largest_face_encoding)
        best_match_index = np.argmin(face_distances)

        if matches[best_match_index]:
            name = known_faces_names[best_match_index]
        else:
            name = "Unrecognized"

        face_names.append(name)

        top, right, bottom, left = largest_face_location

        # Масштабирование координат до оригинального размера
        top *= 4
        right *= 4
        bottom *= 4
        left *= 4

        # Предсказание наличия СИЗ на кадре
        results = model(frame)
        has_hardhat = has_vest = False

        for result in results:
            boxes = result.boxes
            for box in boxes:
                x1, y1, x2, y2 = map(int, box.xyxy[0])
                width, height = x2 - x1, y2 - y1

                if width < CFG.MIN_BOX_SIZE or height < CFG.MIN_BOX_SIZE:
                    continue

                cls_id = int(box.cls[0])
                label = model.names[cls_id]

                if label == "Person" or label == "Mask" or label == "NO-Mask":
                    continue

                item_color = (0, 255, 0) if "NO" not in label else (0, 0, 255)

                if label == "Hardhat":
                    has_hardhat = True
                elif label == "Safety Vest":
                    has_vest = True

                cv2.rectangle(frame, (x1, y1), (x2, y2), item_color, 2)
                cv2.putText(frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, item_color, 2, cv2.LINE_AA)

        # Запись данных о человеке и СИЗ
        box_color = (0, 255, 0) if has_hardhat and has_vest else (0, 165, 255) if has_hardhat or has_vest else (0, 0, 255)
        cv2.rectangle(frame, (left, top), (right, bottom), box_color, 2)
        cv2.rectangle(frame, (left, bottom - 35), (right, bottom), box_color, cv2.FILLED)
        cv2.putText(frame, name, (left + 6, bottom - 6), cv2.FONT_HERSHEY_DUPLEX, 0.8, (255, 255, 255), 1)

        now = datetime.now()
        if (now - last_full_record_time) > full_log_interval:
            last_full_record_time = now
            ws.append([now.strftime("%H:%M:%S"), name, 'Yes' if has_hardhat else 'No', 'Yes' if has_vest else 'No'])
            wb.save(excel_file)

    # Отображение кадра в интерфейсе
    img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    img = Image.fromarray(img)
    imgtk = ImageTk.PhotoImage(image=img)
    video_label.imgtk = imgtk
    video_label.configure(image=imgtk)

    video_label.after(10, update_frame)

# Создание графического интерфейса
root = tk.Tk()
root.title("PPE Detection System")

root.configure(background="#B2DFDB")

# Добавление кнопки для открытия Excel файла
open_button = tk.Button(root, text="Open Excel File", command=open_excel_file, font="helvetica 13", foreground="#004D40", background="#B2DFDB")
open_button.pack(pady=20, padx=10)

# Добавление виджета для видео
video_label = tk.Label(root)
video_label.pack()

# Запуск видеокамеры
video_capture = cv2.VideoCapture(0)
update_frame()

# Запуск интерфейса
root.mainloop()